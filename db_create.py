################################################################################
#  2810ict - Assignment 2
#
#
# Access workbooks and create a database
#
#
# Created by Brianna Sonter | s2930629
#
################################################################################

#Imports
import re
import sqlite3
import openpyxl
from openpyxl import *
import sys

#Create/connect to databse
def connectDB():
    try:
        db_fname = "data/foodData.db"
        connection = sqlite3.connect(db_fname)
        #let user know that connection to the database was successful
        print("Connected to Database successfully")
        return connection

    except:
        #let user know that connection to database couldnt be created
        print("Error: Could not connect to Database")

#function to close the connection to database
def closeDB(_connection):
    try:
        connection = _connection
        connection.close()
        print("Connection closed")

    except:
        #let user know database was unable to close
        print("Error: Was unable to close connection to Database")

#Open database
connection = connectDB()
cursor = connection.cursor()

#function to create the tables in the database
def createFoodDatabase():
    try:
        #create table for inspections
        create_inspections_table = """ CREATE TABLE IF NOT EXISTS inspections (
        activity_date DATETIME,
        employee_id TEXT,
        facility_address TEXT,
        facility_city TEXT,
        facility_id TEXT,
        facility_name TEXT,
        facility_state TEXT,
        facility_zip TEXT,
        grade TEXT,
        owner_id TEXT,
        owner_name TEXT,
        pe_description TEXT,
        program_element_pe NUMERIC,
        program_name TEXT,
        program_status TEXT,
        record_id TEXT,
        score NUMERIC,
        serial_number TEXT,
        service_code NUMERIC,
        service_description TEXT);
        """
        cursor.execute(create_inspections_table)

        #create table for violations
        create_violations_table = """ CREATE TABLE IF NOT EXISTS violations (
        points NUMBER(1),
        serial_number VARCHAR(9),
        violation_code CHAR(4),
        violation_description VARCHAR(100),
        violation_status VARCHAR(20)
        );
        """
        cursor.execute(create_violations_table)

        print("Database table creation was successful")

    except:
        print("Error: Database table creation was a failure :(")
        closeDB(connection)

#function to make excel data ready for transfer into the sqlite database
def inspections():
    fname = "data/inspections.xlsx"

    #prompt to let user know data is loading
    sys.stdout.write("Loading inspections spreadsheet ... ")
    sys.stdout.flush()

    #try to open from excel
    try:
        workbook_inspect = load_workbook(fname)
        sheet_ranges_inspect = workbook_inspect.sheetnames
        sheet_inspect = workbook_inspect[sheet_ranges_inspect[0]]
        print("sucessful data load from inspections excel")
        return sheet_inspect

    except:
        print("Error: Could not load data from inspections spreadsheet")
        closeDB(connection)

#function to make excel data ready for transfer into the sqlite database
def violations():
    fname_violations = "data/violations.xlsx"

    #prompt to let user know data is loading
    sys.stdout.write("Loading violations spreadsheet ... ")
    sys.stdout.flush()

    #try to open from excel
    try:
        workbook_violations = load_workbook(fname_violations)
        sheet_ranges_violations = workbook_violations.sheetnames
        sheet_violations = workbook_violations[sheet_ranges_violations[0]]

        print("sucessful data load from violations excel")
        return sheet_violations

    except:
        print("Error: Could not load data from violations spreadsheet")
        closeDB(connection)

#function to insert data from excel into sqlite database
def insertInspections(_worksheet):
     #prompt to let user know data is being input
    sys.stdout.write("Loading data from inspections spreadsheet ... ")
    sys.stdout.flush()

    #locate data for each section of database
    for row in _worksheet:
        _activity_date = row[0].value
        _employee_id = row[1].value
        _facility_address = row[2].value
        _facility_city = row[3].value
        _facility_id = row[4].value
        _facility_name = row[5].value
        _facility_state = row[6].value
        _facility_zip = row[7].value
        _grade = row[8].value
        _owner_id = row[9].value
        _owner_name = row[10].value
        _pe_description = row[11].value
        _program_element_pe = row[12].value
        _program_name = row[13].value
        _program_status = row[14].value
        _record_id = row[15].value
        _score = row[16].value
        _serial_number = row[17].value
        _service_code = row[18].value
        _service_description = row[19].value

        # get ready to insert into sqlite databse table previously created
        data = """INSERT INTO inspections (
        activity_date,
        employee_id,
        facility_address,
        facility_city,
        facility_id,
        facility_name,
        facility_state,
        facility_zip, grade,
        owner_id, owner_name,
        pe_description,
        program_element_pe,
        program_name,
        program_status,
        record_id,
        score,
        serial_number,
        service_code,
        service_description)
        VALUES ( ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?);
        """
        values = (_activity_date, _employee_id, _facility_address, _facility_city, _facility_id, _facility_name, _facility_state, _facility_zip, _grade, _owner_id, _owner_name, _pe_description, _program_element_pe, _program_name, _program_status, _record_id, _score, _serial_number, _service_code, _service_description)
            #print(sql)

        #match values with how to insert it and insert into database
        cursor.execute(data, values)

    #commit data to database
    connection.commit()
    #let user know data has been commited
    print("Loaded data successfully")

#function to insert data from excel into sqlite database
def insertViolations(_worksheet):
    #prompt to let user know that data is being loaded into database
    sys.stdout.write("Loading data from violations spreadsheet ... ")
    sys.stdout.flush()

    #locate data for each section of the database
    for row in _worksheet:
        _points = row[0].value
        _serial_number = row[1].value
        _violation_code = row[2].value
        _violation_description = row[3].value
        _violation_status = row[4].value

        data = """
        INSERT INTO violations(
        points,
        serial_number,
        violation_code,
        violation_description,
        violation_status
        )
        VALUES( ?, ?, ?, ?, ?)
        """

        values = (_points, _serial_number, _violation_code, _violation_description, _violation_status)

        #match values with how to insert data and insert into database
        cursor.execute(data, values)

    #commit data to database
    connection.commit()
    #let user know data has been commited
    print("Data was inserted successfully for violations")


#run funtions to create database
if __name__ == '__main__':
    createFoodDatabase()

    inspect_data = inspections()
    insertInspections(inspect_data)

    violation_data = violations()
    insertViolations(violation_data)

    closeDB(connection)
